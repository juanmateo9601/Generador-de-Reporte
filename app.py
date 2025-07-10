import streamlit as st
import pandas as pd
import io, csv
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import range_boundaries, get_column_letter
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
import os # Para manejar rutas de archivos

import unicodedata

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return texto
    try:
        # Decodificar errores comunes mal interpretados desde latin1
        texto = texto.encode("latin1").decode("utf-8")
    except:
        pass
    # Eliminar caracteres no imprimibles si quedan
    return unicodedata.normalize("NFKC", texto)


st.set_page_config(page_title="Generador Reporte Final", layout="centered")
st.title("📄 Generador de Reporte")
st.markdown("Sube el archivo CSV exportado y generaremos el Excel automáticamente.")

plantilla_seleccionada = st.selectbox(
    "📑 Selecciona la plantilla de destino",
    options=["Medellín", "Findeter"],
    index=0
)

plantillas = {
    "Medellín": "Plantilla_Medellin_Final.xlsx",
    "Findeter": "Plantilla_Findeter_Final.xlsx"
}

plantilla_path = plantillas[plantilla_seleccionada]


# EXTRAER DATOS DESDE TEXTO (BENEFICIARIO Y TÉCNICO)
def extraer_datos_desde_texto(archivo_subido):
    archivo_subido.seek(0)
    contenido = archivo_subido.read().decode("latin1").splitlines()

    datos = {
        "nombre": "", "cedula": "", "direccion": "", "telefono": "", "telefono2": "", "idhogar": "",
        "tecnico_nombre": "", "tecnico_cedula": "", "tecnico_cargo": ""
    }

    print("\n📥 Procesando archivo línea por línea...")

    for linea in contenido:
        linea_original = linea
        linea = (
            linea.replace("CÃ©dula", "Cédula")
                 .replace("TelÃ©fono", "Teléfono")
                 .replace("DirecciÃ³n", "Dirección")
                 .replace("TÃ©cnico", "Técnico")
                 .replace("INFORMACIÃN DEL TÃCNICO", "INFORMACIÓN DEL TÉCNICO")
        )

        print("▶", linea)

        if ":" in linea:
            partes = linea.split(":", 1)
            clave = partes[0].strip().lower()
            valor = normalizar_texto(partes[1].replace('"', '').strip().lstrip(','))



            if "nombre técnico" in clave:
                datos["tecnico_nombre"] = valor
            elif "cédula técnico" in clave:
                datos["tecnico_cedula"] = valor
            elif "cargo técnico" in clave:
                datos["tecnico_cargo"] = valor
            elif "nombre" in clave and "técnico" not in clave:
                datos["nombre"] = valor
            elif "cédula" in clave and "técnico" not in clave:
                datos["cedula"] = valor
            elif "dirección" in clave:
                datos["direccion"] = valor
            elif "teléfono 1" in clave:
                datos["telefono"] = valor
            elif "teléfono 2" in clave:
                datos["telefono2"] = valor
            elif "id hogar" in clave:
                datos["idhogar"] = valor

    print("✅ Datos extraídos:", datos)
    st.write("📋 Datos extraídos desde texto:", datos)
    return datos, contenido

def nombre_a_archivo(nombre):
    nombre = nombre.strip().lower()
    nombre = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode("utf-8")
    nombre = nombre.replace(" ", "_")
    return f"{nombre}.png"


# EXTRAER ACTIVIDADES DESDE TEXTO
def extraer_actividades_desde_texto(lineas):
    idx_inicio = next((i for i, l in enumerate(lineas) if l.lower().startswith("item,")), None)
    idx_fin = next((i for i, l in enumerate(lineas) if "TOTAL GENERAL" in l.upper()), None)

    if idx_inicio is None:
        st.error("❌ No se encontró encabezado de actividades.")
        return pd.DataFrame()

    actividades_raw = lineas[idx_inicio:idx_fin]

    # Convertimos a DataFrame
    actividades_csv = "\n".join(actividades_raw)
    from io import StringIO
    df = pd.read_csv(StringIO(actividades_csv), encoding="latin1")

    # Limpiar signos raros en texto de todas las columnas tipo string
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].apply(normalizar_texto)


    # Normalizamos columna Categoría si está mal codificada
    if "CategorÃ­a" in df.columns:
        df.rename(columns={"CategorÃ­a": "Categoría"}, inplace=True)

    return df

# FUNCIÓN PARA ESCRIBIR EN CELDA (CONCATENANDO)
def escribir_en_celda(ws, celda_destino, valor, formato=None):
    for rango in ws.merged_cells.ranges:
        if celda_destino in rango:
            min_col, min_row, *_ = range_boundaries(str(rango))
            celda_destino = f"{get_column_letter(min_col)}{min_row}"
            break

    valor_original = ws[celda_destino].value or ""
    if valor_original and str(valor).strip() not in str(valor_original):
        nuevo_valor = f"{valor_original} {valor}".strip()
    else:
        nuevo_valor = valor_original or valor

    ws[celda_destino] = nuevo_valor

    col = celda_destino[0].upper()
    alineacion = Alignment(horizontal="left", wrap_text=True)
    fuente = Font(name="Times New Roman", size=14, color="000000")  # negro


    ws[celda_destino].alignment = alineacion
    ws[celda_destino].font = fuente

    if formato:
        ws[celda_destino].number_format = formato

    print(f"🖊️ Escribiendo en {celda_destino}: '{nuevo_valor}'")




def ajustar_altura_fila(ws, fila, col='C'):
    celda = f"{col}{fila}"
    valor = str(ws[celda].value) if ws[celda].value else ""
    if not valor:
        return
    ancho_col = ws.column_dimensions[col].width or 100
    lineas = sum([len(line) // int(ancho_col) + 1 for line in valor.split('\n')])
    ws.row_dimensions[fila].height = max(15, 15 * lineas)

def set_print_area(ws, col_inicio="A", col_fin="G", fila_inicio=1, fila_fin=None):
    if fila_fin is None:
        fila_fin = ws.max_row
    ws.print_area = f"{col_inicio}{fila_inicio}:{col_fin}{fila_fin}"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

# FUNCIÓN PARA ESCRIBIR LA PLANTILLA FINAL
def escribir_plantilla(df, datos, plantilla_path, tipo_plantilla):
    wb = load_workbook(plantilla_path)
    ws = wb.active

    # === REORDENAR ACTIVIDADES SEGÚN ARCHIVO DE REFERENCIA ===
    try:
        df["Item"] = df["Item"].astype(str)
        orden_referencia = pd.read_excel("MEDELLIN_ARCHIVO_PARA_TRABAJAR.xlsx", sheet_name="FORMATO DE OFERTA ECONÓMICA")
        orden_referencia = orden_referencia[orden_referencia["Item"].notna()]
        orden_referencia["Item"] = orden_referencia["Item"].astype(str)
        df = df.merge(orden_referencia[["Item"]].reset_index().rename(columns={"index": "orden_idx"}), on="Item", how="left")
        df["orden_idx"] = df["orden_idx"].fillna(999999).astype(int)
        df = df.sort_values(by="orden_idx")
        print("✅ Actividades ordenadas según archivo de referencia.")
    except Exception as e:
        print(f"⚠️ No se pudo ordenar según archivo de referencia: {e}")

    print(f"\n🧾 Inyectando datos para plantilla: {tipo_plantilla}")
    ws.column_dimensions['C'].width = 60

    if tipo_plantilla == "Medellín":
        campos = {
            "C7": datos.get("nombre", ""),
            "C8": datos.get("cedula", ""),
            "C9": f'{datos.get("telefono", "")} / {datos.get("telefono2", "")}'.strip(" /"),
            "D8": datos.get("direccion", ""),
            "G7": datos.get("idhogar", ""),
            "G4": datetime.now().strftime("%d/%m/%Y"),
            "G9": datetime.now().strftime("%d/%m/%Y")
        }
        fila_inicio = 14
        fila_totales = {
            "subtotal": "G77",
            "iva": "G81",
            "aiu": "G82",
            "total": "G83",
            "valor_final": "G85"
        }
        celda_tecnico_nombre = "B100"
        celda_tecnico_cedula = "C101"
    elif tipo_plantilla == "Findeter":
        campos = {
            "F15": datos.get("nombre", ""),
            "E16": datos.get("cedula", ""),
            "F17": f'{datos.get("telefono", "")} / {datos.get("telefono2", "")}'.strip(" /"),
            "B16": datos.get("direccion", ""),
            "B17": datos.get("cedula", ""),
            "G5": datos.get("idhogar", ""),
            "G6": datetime.now().strftime("%d/%m/%Y")
        }
        fila_inicio = 31
        fila_totales = {
            "subtotal": "G93",
            "valor_final": "G94"
        }
        celda_tecnico_nombre = "B104"
        celda_tecnico_cedula = "B105"

    for celda, valor in campos.items():
        escribir_en_celda(ws, celda, valor)

    celdas_no_editables = set()
    for r in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(r))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                celdas_no_editables.add(f"{get_column_letter(col)}{row}")

    categorias = df["Categoría"].dropna().unique()

    for cat in categorias:
        if f"B{fila_inicio}" not in celdas_no_editables:
            item_categoria = (
                df[df["Categoría"] == cat]["Item"]
                .astype(str)
                .str.extract(r'^(\d{1,3})')[0]
                .dropna()
                .iloc[0]
                if not df[df["Categoría"] == cat].empty else ""
            )

            ws[f"B{fila_inicio}"] = item_categoria
            ws[f"C{fila_inicio}"] = cat
            ws[f"B{fila_inicio}"].font = Font(bold=True)
            ws[f"C{fila_inicio}"].font = Font(bold=True)
            for col in ["B", "C", "D", "E", "F", "G"]:
                ws[f"{col}{fila_inicio}"].fill = PatternFill("solid", fgColor="D3D3D3")

        fila_inicio += 1
        fila_ini_actividades = fila_inicio

        actividades = df[df["Categoría"] == cat]
        for _, row in actividades.iterrows():
            def limpiar_valor_moneda(valor):
                if isinstance(valor, str):
                    valor = valor.replace("$", "").replace(".", "").replace(",", ".")
                try:
                    return int(float(valor) * 100) / 100
                except:
                    return 0.0

            datos_fila = {
                f"B{fila_inicio}": row["Item"],
                f"C{fila_inicio}": normalizar_texto(row["Actividad Obra"]),
                f"D{fila_inicio}": row["Un"],
                f"E{fila_inicio}": float(row["Cant"]),
                f"F{fila_inicio}": limpiar_valor_moneda(row["V. Unitario"]),
                f"G{fila_inicio}": limpiar_valor_moneda(row["V. Parcial"]),
            }

            for celda, valor in datos_fila.items():
                if celda not in celdas_no_editables:
                    ws[celda] = valor
                    ws[celda].alignment = Alignment(horizontal="left", wrap_text=True)

                    # Asegurar fuente negra para todos los campos
                    ws[celda].font = Font(color="000000", name="Times New Roman", size=12)

                    # Formato moneda si aplica
                    if celda[0] in ["E", "F", "G"]:
                        ws[celda].number_format = '"$"#,##0.00'


            ajustar_altura_fila(ws, fila_inicio, 'C')
            fila_inicio += 1

        ws[f"F{fila_inicio}"] = "SUBTOTAL"
        ws[f"F{fila_inicio}"].font = Font(bold=True)
        ws[f"G{fila_inicio}"] = f"=SUM(G{fila_ini_actividades}:G{fila_inicio - 1})"
        ws[f"G{fila_inicio}"].font = Font(bold=True)
        ws[f"G{fila_inicio}"].number_format = '"$"#,##0.00'
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{fila_inicio}"].fill = PatternFill("solid", fgColor="DDDDDD")
        fila_inicio += 1

    # === TOTAL GENERAL basado en celdas con "SUBTOTAL" en columna C ===
    # === TOTAL GENERAL basado en celdas con "SUBTOTAL" en columna F ===
    subtotales_reales = []
    for fila in range(1, ws.max_row + 1):
        if str(ws[f"F{fila}"].value).strip().upper() == "SUBTOTAL":
            subtotales_reales.append(f"G{fila}")

    formula_subtotal = f"=SUM({','.join(subtotales_reales)})" if subtotales_reales else "=0"
    escribir_en_celda(ws, fila_totales["subtotal"], formula_subtotal, '"$"#,##0.00')

    # IVA, AIU y Total: también dinámicos
    if tipo_plantilla == "Medellín":
        subtotal_cell = fila_totales["subtotal"]
        iva_cell = fila_totales["iva"]
        aiu_cell = fila_totales["aiu"]
        total_cell = fila_totales["total"]
        valor_final_cell = fila_totales["valor_final"]

        escribir_en_celda(ws, iva_cell, f"={subtotal_cell}*0.12", '"$"#,##0.00')
        escribir_en_celda(ws, aiu_cell, f"={subtotal_cell}*0.016", '"$"#,##0.00')
        escribir_en_celda(ws, total_cell, f"={subtotal_cell}+{iva_cell}+{aiu_cell}", '"$"#,##0.00')
        escribir_en_celda(ws, valor_final_cell, f"={total_cell}", '"$"#,##0.00')
    elif tipo_plantilla == "Findeter":
        escribir_en_celda(ws, fila_totales["valor_final"], f"={fila_totales['subtotal']}", '"$"#,##0.00')


        # Recolectar todos los subtotales válidos
        subtotales_reales = []
        for fila in range(1, ws.max_row + 1):
            if str(ws[f"F{fila}"].value).strip().upper() == "SUBTOTAL":
                subtotales_reales.append(f"G{fila}")

        # Crear fórmula con suma explícita de todos los subtotales (G45 + G52 + ...)
        formula_subtotal = f"=SUM({','.join(subtotales_reales)})" if subtotales_reales else "=0"
        escribir_en_celda(ws, fila_totales["subtotal"], formula_subtotal, '"$"#,##0.00')

        escribir_en_celda(ws, fila_totales["iva"], f"={fila_totales['subtotal']}*0.12", '"$"#,##0.00')
        escribir_en_celda(ws, fila_totales["aiu"], f"={fila_totales['subtotal']}*0.016", '"$"#,##0.00')
        escribir_en_celda(ws, fila_totales["total"], f"={fila_totales['subtotal']}+{fila_totales['iva']}+{fila_totales['aiu']}", '"$"#,##0.00')
        escribir_en_celda(ws, fila_totales["valor_final"], f"={fila_totales['total']}", '"$"#,##0.00')

    elif tipo_plantilla == "Findeter":
        escribir_en_celda(ws, fila_totales["valor_final"], f"={fila_totales['subtotal']}", '"$"#,##0.00')

    escribir_en_celda(ws, celda_tecnico_nombre, datos.get("tecnico_nombre", ""))
    escribir_en_celda(ws, celda_tecnico_cedula, datos.get("tecnico_cedula", ""))

    # === OCULTAR FILAS VACÍAS ENTRE CONTENIDO Y TOTALES ===
    fila_totales_inicio = int(fila_totales["subtotal"][1:])  # e.g. G77 -> 77

    for fila in range(fila_inicio, fila_totales_inicio):
        # Solo ocultar si toda la fila está vacía
        if all(ws[f"{col}{fila}"].value in ("", None) for col in ["A", "B", "C", "D", "E", "F", "G"]):
            ws.row_dimensions[fila].hidden = True

    for fila in range(1, ws.max_row + 1):
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            celda = f"{col}{fila}"
            if isinstance(ws[celda], MergedCell):
                continue
            if ws[celda].value is None:
                ws[celda].value = ""

    set_print_area(ws, col_inicio="A", col_fin="G", fila_inicio=1, fila_fin=ws.max_row)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "portrait"

        # === INSERTAR FIRMA EN C93 ===
    firma_nombre = datos.get("tecnico_nombre", "")
    st.write(f"👤 Nombre del técnico extraído: '{firma_nombre}'")

    archivo_firma = nombre_a_archivo(firma_nombre)
    ruta_firma = os.path.join("firmas", archivo_firma)

    st.write(f"🛣️ Buscando firma en: `{ruta_firma}`")

    print(f"🛠️ Preparando inserción de firma para: {firma_nombre}")
    print(f"📁 Ruta esperada: {ruta_firma}")

    if os.path.exists(ruta_firma):
        try:
            img = Image(ruta_firma)
            img.width = 140  # Ajusta según necesidad
            img.height = 85
            ws.add_image(img, "D100")

            print(f"✅ Firma insertada correctamente desde: {ruta_firma}")
            st.success(f"Firma insertada correctamente para **{firma_nombre}**")
        except Exception as e:
            print(f"❌ Error al insertar imagen en Excel: {e}")
            st.error(f"❌ Error al insertar la firma de **{firma_nombre}**: {e}")
    else:
        print(f"⚠️ Archivo de firma NO encontrado: {ruta_firma}")
        st.warning(f"⚠️ No se encontró la firma para el técnico **{firma_nombre}**. Archivo esperado: `{ruta_firma}`")


    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output




# INTERFAZ PRINCIPAL
csv_file = st.file_uploader("📤 Sube el archivo CSV generado", type=["csv"])

if csv_file:
    datos_extraidos, lineas_csv = extraer_datos_desde_texto(csv_file)
    df_actividades = extraer_actividades_desde_texto(lineas_csv)

    if not df_actividades.empty:
        st.write("🧠 Vista previa de actividades:")
        st.dataframe(df_actividades.head(10))

        if st.button("🛠️ Generar Excel"):
            archivo = escribir_plantilla(df_actividades, datos_extraidos, plantilla_path, plantilla_seleccionada)
            st.download_button(
                label="📥 Descargar Excel generado",
                data=archivo,
                file_name=f"Reporte_Medellin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
